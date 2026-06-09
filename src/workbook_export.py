"""Workbook tab extraction, compact context, and patch application."""

from __future__ import annotations

import json
import logging
import re
import shutil
from pathlib import Path
from typing import Any

from src.config import settings

logger = logging.getLogger(__name__)

_PATCH_RE = re.compile(r"^(.+?)!([A-Za-z]+\d+)\s*=\s*(.*)\s*$", re.DOTALL)

WORKBOOK_JSON_FILENAMES: dict[str, str] = {
    "guide_workbook.xlsm": "guide_workbook.json",
    "example_workbook.xlsx": "example_workbook.json",
}


def get_workbook_tab_names(workbook_path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required") from exc

    wb = load_workbook(workbook_path, read_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def default_workbook_template() -> Path:
    return settings.data_dir / "guide_workbook.xlsm"


def parse_workbook_patch(patch: str) -> tuple[str, str, str] | None:
    line = patch.strip()
    if not line or line.startswith("#"):
        return None
    match = _PATCH_RE.match(line)
    if not match:
        return None
    sheet, cell, value = match.group(1).strip(), match.group(2).strip(), match.group(3)
    return sheet, cell.upper(), value


def extract_patches_from_text(text: str) -> list[str]:
    """Extract workbook patches from Claude response markers or fenced blocks."""
    markers = [
        ("--- WORKBOOK PATCHES START ---", "--- WORKBOOK PATCHES END ---"),
        ("```workbook_patches", "```"),
        ("```patches", "```"),
    ]
    for start, end in markers:
        if start in text:
            block = text.split(start, 1)[1]
            if end in block:
                block = block.split(end, 1)[0]
            return [ln for ln in block.strip().splitlines() if ln.strip() and not ln.strip().startswith("#")]

    patches: list[str] = []
    for line in text.splitlines():
        if "!" in line and "=" in line and parse_workbook_patch(line):
            patches.append(line.strip())
    return patches


def validate_patches(
    patches: list[str],
    valid_sheets: list[str],
) -> tuple[list[str], list[str]]:
    """Return (valid_patches, rejection_messages)."""
    valid: list[str] = []
    rejected: list[str] = []
    sheet_set = set(valid_sheets)

    for raw in patches:
        parsed = parse_workbook_patch(raw)
        if not parsed:
            if raw.strip():
                rejected.append(f"unparseable: {raw[:80]}")
            continue
        sheet_name, _, _ = parsed
        if sheet_name not in sheet_set:
            rejected.append(
                f"unknown sheet {sheet_name!r} (valid: {', '.join(valid_sheets)})"
            )
            continue
        valid.append(raw.strip())

    return valid, rejected


def apply_patches_to_workbook(
    patches: list[str],
    output_path: Path,
    *,
    template_path: Path | None = None,
    valid_sheets: list[str] | None = None,
) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required") from exc

    template_path = Path(template_path or default_workbook_template())
    output_path = Path(output_path)
    if output_path.exists():
        raise FileExistsError(f"Refusing to overwrite: {output_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    if valid_sheets is None:
        valid_sheets = get_workbook_tab_names(template_path)

    validated, rejected = validate_patches(patches, valid_sheets)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path, keep_vba=template_path.suffix.lower() == ".xlsm")
    applied = 0

    for raw in validated:
        parsed = parse_workbook_patch(raw)
        if not parsed:
            continue
        sheet_name, coordinate, value = parsed
        ws = wb[sheet_name]
        cell = ws[coordinate]
        if value.startswith("="):
            cell.value = value
        else:
            stripped = value.strip()
            if stripped.lower() in ("true", "false"):
                cell.value = stripped.lower() == "true"
            else:
                try:
                    cell.value = float(stripped) if "." in stripped else int(stripped)
                except ValueError:
                    cell.value = value
        applied += 1

    wb.save(str(output_path))
    logger.info("Workbook saved: %s (%s patches applied)", output_path, applied)

    return {
        "output_path": str(output_path),
        "patches_submitted": len(patches),
        "patches_applied": applied,
        "patches_rejected": len(rejected),
        "rejection_details": rejected[:20],
        "valid_sheets": valid_sheets,
    }


def build_compact_workbook_json(source_path: Path) -> dict[str, Any]:
    """Build compact JSON: labels + formulas only (no numeric outputs)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required") from exc

    source_path = Path(source_path)
    mode = "structure" if source_path.name == "example_workbook.xlsx" else "template"
    wb = load_workbook(source_path, data_only=False, read_only=True)
    sheets: dict[str, list[list[str]]] = {}

    try:
        for sheet in wb.worksheets:
            rows: list[list[str]] = []
            for row in sheet.iter_rows():
                cells: list[str] = []
                for cell in row:
                    text = _compact_cell(cell, mode)
                    if text is not None:
                        cells.append(text)
                if cells:
                    rows.append(cells)
            if rows:
                sheets[sheet.title] = rows
    finally:
        wb.close()

    return {
        "source_file": source_path.name,
        "mode": mode,
        "sheet_count": len(sheets),
        "sheet_names": list(sheets.keys()),
        "sheets": sheets,
    }


def _compact_cell(cell: Any, mode: str) -> str | None:
    value = cell.value
    if value is None:
        return None
    data_type = cell.data_type
    if mode == "structure":
        if data_type != "s":
            return None
        text = str(value).strip()
        return text or None
    if data_type == "n":
        return None
    if data_type == "f":
        return str(value)
    text = str(value).strip()
    return text or None


def compact_workbook_to_prompt_text(data: dict[str, Any]) -> str:
    lines = [
        f"WORKBOOK {data.get('source_file', 'workbook')}",
        f"TABS: {', '.join(data.get('sheet_names', []))}",
    ]
    for sheet_name, rows in data.get("sheets", {}).items():
        lines.append(f"SHEET {sheet_name}")
        for row in rows[:200]:
            lines.append("\t".join(str(c) for c in row))
        if len(rows) > 200:
            lines.append(f"[... {len(rows) - 200} more rows omitted ...]")
    return "\n".join(lines)


def ensure_compact_workbook_json(source_path: Path, json_path: Path) -> dict[str, Any]:
    source_path = Path(source_path)
    json_path = Path(json_path)
    rebuild = not json_path.exists()
    if json_path.exists():
        try:
            rebuild = json_path.stat().st_mtime < source_path.stat().st_mtime
        except OSError:
            rebuild = True
    if rebuild:
        data = build_compact_workbook_json(source_path)
        json_path.parent.mkdir(parents=True, exist_ok=True)
        json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return data
    return json.loads(json_path.read_text(encoding="utf-8"))


def workbook_context_for_prompt(workbook_path: Path, mode: str | None = None) -> str:
    """Return workbook context sized for LLM prompts."""
    mode = (mode or settings.reference_workbook_mode).lower()
    workbook_path = Path(workbook_path)

    if mode == "files_api":
        tabs = get_workbook_tab_names(workbook_path)
        return (
            f"Workbook file attached: {workbook_path.name}\n"
            f"Tab names ({len(tabs)}): {', '.join(tabs)}\n"
            "Use only these exact tab names for any cell patches."
        )

    if mode in ("compact_extract", "compact"):
        json_path = settings.data_dir / WORKBOOK_JSON_FILENAMES.get(
            workbook_path.name, f"{workbook_path.stem}.json"
        )
        data = ensure_compact_workbook_json(workbook_path, json_path)
        return compact_workbook_to_prompt_text(data)

    from src.document_ingest import extract_workbook_structure

    tabs = get_workbook_tab_names(workbook_path)
    structure = extract_workbook_structure(
        workbook_path.read_bytes(), max_rows_per_sheet=80
    )
    return f"TABS: {', '.join(tabs)}\n\n{structure}"
